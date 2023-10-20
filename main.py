import os
# from tabulate import tabulate
import openpyxl
import numpy as np
import pandas as pd
import camelot
import PyPDF2
from tqdm import tqdm
pdf_path = "каталог 1.2 Изм 12.pdf"

def pdf_to_tab(patch, pages, put_out):
    # pages = ['5']
    rasp = {}
    exel_nam = f'{put_out}.xlsx'
    raspozn = 0
    lis = len(pages)
    stran = 0
    for pag in tqdm(pages):
        tab = 0
        nerasp = 2
        # print(rasp)
        try:
            tables = camelot.read_pdf(patch, pages=pag)
            # print(tabulate(tables))
            pd.set_option('display.max_columns', None)
            if tables.n > 1:
                tab = 2
                par_st = 0
            for g in range(0, len(tables)):
                if tab == 2:
                    par_st += 1
                    page = pag + '_' + str(par_st)
                else:
                    page = pag

                df1 = tables[g].df
                # print('df1\n',df1)
                no = np.column_stack([df1[col].str.contains(r"Наименование \nобъекта \nстроительства", na=False) for col in df1])
                e, t = np.where((no==True))
                ved = np.column_stack([df1[col].str.contains(r"Ведомость", na=False) for col in df1])
                q, w = np.where((ved==True))
                metod = 0
                if len(e) == 0 and len(q) == 0:
                    mask0 = np.column_stack([df1[col].str.contains(r"\nНаименование\n", na=False) for col in df1])
                    st, col = np.where(mask0 == True)
                    metod = 0
                    if len(st)==0:
                        mask = np.column_stack([df1[col].str.contains(r"Наименование", na=False) for col in df1])
                        st, col = np.where(mask == True)
                        metod = 1
                    # print('metod', metod)
                    if metod == 1:
                        if len(st) > 0:
                            cols = df1[col[0]].tolist()
                            sst = st[0]-1
                            for i in cols[st[0]:]:
                                sst += 1
                                if 'Заказчик:' in i:
                                    break
                            df2 = df1.iloc[st[0]:sst, col[0]:]

                            df2.columns = df2.iloc[0]
                            df2 = df2.drop(df2.index[0])
                            # print(df2)
                            df2 = df2[df2['Наименование'].str.len() > 0]
                            # print('1')
                            df2 = df2.reset_index().drop('index', axis=1)
                            # print(df2)
                            if len(df2.index) > 0:
                                nam = df2.iloc[0,0]
                                new_df = df2['Наименование'].str.split(' ', expand=True)
                                # print('2')
                                df3 = pd.concat([new_df, df2],axis=1)
                                df3 = df3.drop('Наименование', axis=1)
                                # print('3')
                                # print('df3\n',df3)
                                df3 = df3.dropna(axis = 1, thresh=0)
                                if os.path.isfile(exel_nam):
                                    with pd.ExcelWriter(exel_nam, mode='a') as writer:
                                        df3.to_excel(writer, startrow=3, index=False, sheet_name=page)
                                else:
                                    with pd.ExcelWriter(exel_nam, mode='w') as writer:
                                        df3.to_excel(writer, startrow=3, index=False, sheet_name=page)

                                wb = openpyxl.load_workbook(filename=exel_nam)
                                ws = wb[page]
                                ws['B2'] =str(nam)
                                wb.save(exel_nam)
                                rasp[f'Страница: {page}'] = 'Распознано'
                                raspozn += 1
                                nerasp = 0
                            else:
                                rasp[f'Страница: {page}'] = 'НЕ РАСПОЗНАНО'
                        else:
                            rasp[f'Страница: {page}'] = 'НЕ РАСПОЗНАНО'
                    if metod == 0:
                        # print('st',len(st))
                        if len(st) > 0:
                            df8 = df1.iloc[st[0]:, col[0]:]
                            # print(df8)
                            new_d = df8[col[0]].str.split('\n', expand=True)
                            df7 = pd.concat([new_d, df1], axis=1)
                            df1 = df7.drop(col[0], axis=1)
                            df1.columns = df1.iloc[0]
                            df1 = df1.drop(df1.index[0])
                            df1 = df1.fillna(value='')
                            col_l = list(df1.columns)
                            # print(col_l)
                            col_i = col_l.index('Наименование')
                            # print((col_i))
                            # print(df1)
                            # print('uyutfdrt')
                            # mask1 = np.column_stack([df1[col].str.contains(r"Наименование", na=False) for col in df1])
                            # print(mask1)
                            # print('ytrer')
                            # st, col = np.where(mask1 == True)
                            # print('e67890-')
                            # if len(st) > 0:
                            #     cols = df1[col[0]].tolist()
                            #     sst = st[0] - 1
                            #     print(cols, sst)
                            #     for i in cols[st[0]:]:
                            #         sst += 1
                            #         if 'Заказчик:' in i:
                            #             break
                            #     df2 = df1.iloc[st[0]:sst, col[0]:]
                            #
                            #     df2.columns = df2.iloc[0]
                            #     df2 = df2.drop(df2.index[0])
                            #     print('df2\n\n',df2)
                            #     df2 = df2[df2['Наименование'].str.len() > 0]
                            #     # print('1')
                            #     df2 = df2.reset_index().drop('index', axis=1)
                            #     # print(df2)
                            if len(df1.index) > 0:
                                nam = df1.iloc[0, 2]
                                # new_df = df1['Наименование'].str.split(' ', expand=True)
                                # # print('2')
                                # df3 = pd.concat([new_df, df1], axis=1)
                                # df3 = df3.drop('Наименование', axis=1)
                                # print('3')
                                # print('df3\n',df3)
                                df2 = df1.iloc[:, col_i:]
                                df3 = df2.dropna(axis=1, thresh=0)
                                df3 = df3.dropna(axis=0, thresh=0)
                                # print('jiuhi')
                                if os.path.isfile(exel_nam):

                                    with pd.ExcelWriter(exel_nam, mode='a') as writer:
                                        df3.to_excel(writer, startrow=3, index=False, sheet_name=page)
                                else:
                                    with pd.ExcelWriter(exel_nam, mode='w') as writer:
                                        df3.to_excel(writer, startrow=3, index=False, sheet_name=page)

                                wb = openpyxl.load_workbook(filename=exel_nam)
                                ws = wb[page]
                                ws['B2'] = str(nam)
                                wb.save(exel_nam)
                                rasp[f'Страница: {page}'] = 'Распознано'
                                raspozn += 1
                                nerasp = 0
                            else:
                                rasp[f'Страница: {page}'] = 'НЕ РАСПОЗНАНО'
                        else:
                            rasp[f'Страница: {page}'] = 'НЕ РАСПОЗНАНО'
                else:
                    rasp[f'Страница: {page}'] = 'НЕ РАСПОЗНАНО'
        except:
            stran += 1
            # print('ERROR')
            if stran == lis:
                exel_nam1 = exel_nam + ' ФАЙЛ НЕ РАСПОЗНАН.txt'
                open(exel_nam1, "w+")
                nerasp = 1
            continue
    ras = pd.DataFrame.from_dict(rasp, orient='index')
    print(ras)
    if raspozn > 0:
        with pd.ExcelWriter(exel_nam, mode='a') as writer:
            ras.to_excel(writer, sheet_name='Распознано')
    else:
        if nerasp >= 1 :
            exel_nam1 = exel_nam + ' ФАЙЛ НЕ РАСПОЗНАН.txt'
            open(exel_nam1, "w+")

def pdf_ex(put_in, put_out):
    try:
        pdfReader = PyPDF2.PdfFileReader(put_in)
        totalPages = pdfReader.numPages
        my_list = [str(i) for i in range(1, totalPages+1)]
        pdf_to_tab(put_in, my_list, put_out)
    except:
        exel_nam = put_out + ' ФАЙЛ НЕ РАСПОЗНАН.txt'
        open(exel_nam, "w+")

# pdf_ex('чертежи/Чертежи для Алрино/ВСб/1ВСб-607-(1-5)-АБ1.pdf','1ВСб-607-(1-5)-АБ1')



# pdf_ex('чертежи/Чертежи для Алрино/Балки Б2.1-2.8/Балка Б2.2-1..pdf','Балка Б2.2-1')